def modify_sheet(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    https://gist.github.com/fredpedroso/590e54d4f07d0ae2d20d0ec0b190d5ff
    
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a',if_sheet_exists='overlay')


    try:
        # try to open an existing workbook
        writer.workbook = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.workbook.sheetnames:
            startrow = writer.workbook[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.workbook.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.workbook.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.workbook.remove(writer.workbook.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.workbook.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.workbook.sheets = {ws.title:ws for ws in writer.workbook.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.close()
